# -*- coding: utf-8 -*-
from __future__ import unicode_literals

import openpyxl

from accounts.models import User
from reviews.forms import DistributionListForm


def import_lists(filepath, category):
    """Import distribution lists from an excel file."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Extracts the user list from the header row
    user_emails = _extract_users(ws)
    user_qs = User.objects.filter(email__in=user_emails)
    users = list(user_qs)
    # Let's make sure we keep user order
    users.sort(key=lambda user: user_emails.index(user.email))

    max_col = len(users) + 1  # Don't use ws.max_column, it's not reliable
    rows = ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=max_col)
    for row in rows:
        _import_list(row, users, category)


def _extract_users(ws):
    """Extract the xls header, i.e the list of email users.

    We cannot rely on the worksheet's dimensions and, say, extract the first
    row until the latest column because this value is sometimes off, especially
    with documents created with LibreOffice.

    """
    users = []
    column_index = 2
    user_cell = ws.cell(row=1, column=column_index)
    while user_cell.value:
        users.append(user_cell.value)
        column_index += 1
        user_cell = ws.cell(row=1, column=column_index)

    return users


def _import_list(row, users, category):
    """Saves a distribution list for a single row."""
    list_name = row[0].value
    reviewers = []
    leader = None
    approver = None

    for idx, cell in enumerate(row[1:]):
        role = cell.value
        if role:
            user = users[idx]

            if role == 'R':
                reviewers.append(user)
            elif role == 'L':
                leader = user
            elif role == 'A':
                approver = user

    data = {
        'name': list_name,
        'categories': [category.id],
        'reviewers': [r.id for r in reviewers],
        'leader': leader.id if leader else None,
        'approver': approver.id if approver else None,
    }
    form = DistributionListForm(data)
    if form.is_valid():
        distribution_list = form.save()
        return distribution_list
