# -*- coding: utf-8 -*-


import os
from io import BytesIO

from django.test import TestCase

import openpyxl

from categories.factories import CategoryFactory
from default_documents.tests.test import ContractorDeliverableTestCase
from accounts.factories import UserFactory
from distriblists.utils import (import_lists, export_lists,
                                export_review_members, import_review_members)
from distriblists.factories import DistributionListFactory
from distriblists.models import DistributionList


class DistributionListsExportTests(TestCase):
    def setUp(self):
        self.category = CategoryFactory()
        self.users = [
            UserFactory(email='user000@test.com', category=self.category),
            UserFactory(email='user001@test.com', category=self.category),
            UserFactory(email='user002@test.com', category=self.category),
            UserFactory(email='user003@test.com', category=self.category),
            UserFactory(email='user004@test.com', category=self.category),
        ]
        no_factory_reviewers = {
            'reviewer1': None,
            'reviewer2': None,
            'reviewer3': None,
        }
        self.lists = [
            DistributionListFactory(
                categories=[self.category],
                leader=self.users[0],
                approver=self.users[1],
                reviewers=[self.users[2], self.users[3]],
                **no_factory_reviewers
            ),
            DistributionListFactory(
                categories=[self.category],
                leader=self.users[2],
                approver=None,
                **no_factory_reviewers
            ),
            DistributionListFactory(
                categories=[self.category],
                leader=self.users[2],
                approver=None,
                reviewers=[self.users[0], self.users[3]],
                **no_factory_reviewers
            ),
        ]

    def test_successful_export(self):
        exported_file = export_lists(self.category)
        buf = BytesIO(exported_file)
        wb = openpyxl.load_workbook(buf)
        ws = wb.active

        self.assertEqual(ws.max_column, 6)

        # Nb rows = total lists + 1
        self.assertEqual(ws.max_row, 4)

        # First row
        self.assertEqual(ws.cell(column=2, row=2).value, 'L')
        self.assertEqual(ws.cell(column=3, row=2).value, 'A')
        self.assertEqual(ws.cell(column=4, row=2).value, 'R')
        self.assertEqual(ws.cell(column=5, row=2).value, 'R')
        self.assertEqual(ws.cell(column=6, row=2).value, None)
        self.assertEqual(ws.cell(column=7, row=2).value, None)

        # Second row
        self.assertEqual(ws.cell(column=2, row=3).value, None)
        self.assertEqual(ws.cell(column=3, row=3).value, None)
        self.assertEqual(ws.cell(column=4, row=3).value, 'L')
        self.assertEqual(ws.cell(column=5, row=3).value, None)
        self.assertEqual(ws.cell(column=6, row=3).value, None)
        self.assertEqual(ws.cell(column=7, row=3).value, None)

        # Third row
        self.assertEqual(ws.cell(column=2, row=4).value, 'R')
        self.assertEqual(ws.cell(column=3, row=4).value, None)
        self.assertEqual(ws.cell(column=4, row=4).value, 'L')
        self.assertEqual(ws.cell(column=5, row=4).value, 'R')
        self.assertEqual(ws.cell(column=6, row=4).value, None)
        self.assertEqual(ws.cell(column=7, row=4).value, None)


class DistributionListsImportTests(TestCase):
    def setUp(self):
        self.category = CategoryFactory()
        self.users = [
            UserFactory(email='user1@test.com', category=self.category),
            UserFactory(email='user2@test.com', category=self.category),
            UserFactory(email='user3@test.com', category=self.category),
            UserFactory(email='user4@test.com', category=self.category),
            UserFactory(email='user5@test.com', category=self.category),
        ]
        UserFactory(email='user6@test.com')

        self.other_category = CategoryFactory()
        self.other_category.users.add(*self.users)

    def test_successful_import(self):
        """Importing the file creates the distribution lists."""
        qs = DistributionList.objects.all()
        self.assertEqual(qs.count(), 0)

        xls_file = os.path.join(
            os.path.dirname(__file__),
            'fixtures',
            'valid_distrib_list.xlsx')
        import_lists(xls_file, self.category)

        self.assertEqual(qs.count(), 5)

        self.assertEqual(qs[0].name, 'Liste 1')
        self.assertEqual(qs[0].leader.email, 'user3@test.com')
        self.assertEqual(qs[0].approver.email, 'user5@test.com')
        self.assertEqual(qs[0].reviewers.count(), 2)

        self.assertEqual(qs[2].name, 'Liste 3')
        self.assertEqual(qs[2].leader.email, 'user5@test.com')
        self.assertIsNone(qs[2].approver)

    def test_importing_twice_with_different_categories(self):
        qs = DistributionList.objects.all()
        self.assertEqual(qs.count(), 0)

        xls_file = os.path.join(
            os.path.dirname(__file__),
            'fixtures',
            'valid_distrib_list.xlsx')
        import_lists(xls_file, self.category)
        import_lists(xls_file, self.other_category)

        self.assertEqual(qs.count(), 5)
        self.assertEqual(len(qs[0].categories.all()), 2)

    def test_import_overrides_existing_content(self):
        """Importing is an idempotent action (PUT, not POST)."""
        qs = DistributionList.objects.all()
        self.assertEqual(qs.count(), 0)

        xls_file = os.path.join(
            os.path.dirname(__file__),
            'fixtures',
            'valid_distrib_list.xlsx')
        import_lists(xls_file, self.category)

        self.assertEqual(qs.count(), 5)

        import_lists(xls_file, self.category)
        self.assertEqual(qs.count(), 5)

        qs[0].delete()
        self.assertEqual(qs.count(), 4)

        import_lists(xls_file, self.category)
        self.assertEqual(qs.count(), 5)

    def test_invalid_data_no_leader(self):
        """Cannot import lists without a leader."""
        qs = DistributionList.objects.all()
        self.assertEqual(qs.count(), 0)

        xls_file = os.path.join(
            os.path.dirname(__file__),
            'fixtures',
            'distrib_list_missing_leader.xlsx')
        import_lists(xls_file, self.category)

        self.assertEqual(qs.count(), 1)

    def test_invalid_data_unknown_user(self):
        """Can only create lists with known users."""
        qs = DistributionList.objects.all()
        self.assertEqual(qs.count(), 0)

        xls_file = os.path.join(
            os.path.dirname(__file__),
            'fixtures',
            'distrib_list_unknown_user.xlsx')
        import_lists(xls_file, self.category)

        self.assertEqual(qs.count(), 1)

    def test_invalid_data_invalid_category(self):
        """Users have to belong to the category."""
        qs = DistributionList.objects.all()
        self.assertEqual(qs.count(), 0)

        xls_file = os.path.join(
            os.path.dirname(__file__),
            'fixtures',
            'distrib_list_wrong_category.xlsx')
        import_lists(xls_file, self.category)

        self.assertEqual(qs.count(), 1)


class ReviewMembersExportTests(ContractorDeliverableTestCase):
    def setUp(self):
        super(ReviewMembersExportTests, self).setUp()
        self.users = [
            UserFactory(email='user000@test.com', category=self.category),
            UserFactory(email='user001@test.com', category=self.category),
            UserFactory(email='user002@test.com', category=self.category),
            UserFactory(email='user003@test.com', category=self.category),
            UserFactory(email='user004@test.com', category=self.category),
        ]
        self.docs = [
            self.create_doc(revision={
                'leader': self.users[0]
            }),
            self.create_doc(revision={
                'leader': self.users[1],
                'approver': self.users[0]
            }),
            self.create_doc(revision={
                'leader': self.users[0],
                'approver': self.users[1],
                'reviewers': [self.users[2], self.users[3]]
            }),
            self.create_doc(),
        ]

    def test_successful_export(self):
        exported_file = export_review_members(self.category)
        buf = BytesIO(exported_file)
        wb = openpyxl.load_workbook(buf)
        ws = wb.active

        self.assertEqual(ws.max_column, 7)
        self.assertEqual(ws.max_row, 5)

        self.assertEqual(ws.cell(column=1, row=2).value, self.docs[0].document_number)
        self.assertEqual(ws.cell(column=2, row=2).value, None)
        self.assertEqual(ws.cell(column=3, row=2).value, 'L')
        self.assertEqual(ws.cell(column=4, row=2).value, None)

        self.assertEqual(ws.cell(column=1, row=3).value, self.docs[1].document_number)
        self.assertEqual(ws.cell(column=2, row=3).value, None)
        self.assertEqual(ws.cell(column=3, row=3).value, 'A')
        self.assertEqual(ws.cell(column=4, row=3).value, 'L')


class ReviewMembersImportTests(ContractorDeliverableTestCase):
    def setUp(self):
        super(ReviewMembersImportTests, self).setUp()
        self.users = [
            UserFactory(email='user000@test.com', category=self.category),
            UserFactory(email='user001@test.com', category=self.category),
            UserFactory(email='user002@test.com', category=self.category),
            UserFactory(email='user003@test.com', category=self.category),
            UserFactory(email='user004@test.com', category=self.category),
        ]
        self.docs = [
            self.create_doc(document_key='document0001'),
            self.create_doc(document_key='document0002'),
            self.create_doc(document_key='document0003'),
            self.create_doc(document_key='document0004'),
        ]

    def test_successful_import(self):
        """Importing the file creates the distribution lists."""
        self.assertIsNone(self.docs[0].latest_revision.leader)

        xls_file = os.path.join(
            os.path.dirname(__file__),
            'fixtures',
            'valid_review_members.xlsx')
        import_review_members(xls_file, self.category)
        self.assertEqual(self.docs[0].latest_revision.leader, self.users[0])
        self.assertEqual(self.docs[0].latest_revision.approver, self.users[1])
        self.assertEqual(self.docs[0].latest_revision.reviewers.all().count(), 3)
